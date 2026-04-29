from fastapi import FastAPI, HTTPException, status, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel, field_validator, model_validator
from typing import Optional, List, Dict, Any
import uvicorn
import logging
import io
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
from datetime import datetime
from core import bot_core, user_states
from features.dicts import REMINDER_TYPES, MOODS
from features.f_config import get_db

ADMIN_PASSWORD_HASH = hashlib.sha256("c0rn_5ouP".encode()).hexdigest()

def verify_admin_password(password: str) -> bool:
    """Verify admin password"""
    return hashlib.sha256(password.encode()).hexdigest() == ADMIN_PASSWORD_HASH

app = FastAPI(title="Psychology Bot Web Interface")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="static"), name="static")
web_sessions = {}

class ChatRequest(BaseModel):
    session_id: str
    message: str

class ChatResponse(BaseModel):
    response: str
    success: bool

class MoodEntryRequest(BaseModel):
    session_id: str
    emoji: str
    note: Optional[str] = None

class ReminderRequest(BaseModel):
    session_id: str
    reminder_type: str
    time: str

class SurveySaveRequest(BaseModel):
    session_id: str
    role: str
    survey_key: str
    responses: List[int]
    open_ended_responses: Optional[List[str]] = []

class UserProfileRequest(BaseModel):
    session_id: str
    status: str
    age: int
    field: str
    field_other: Optional[str] = None
    student_course: Optional[int] = None
    student_education_level: Optional[str] = None
    teacher_experience: Optional[int] = None
    employee_experience: Optional[int] = None
    
    @field_validator('age')
    @classmethod
    def validate_age(cls, v: int, info) -> int:
        status  = info.data.get('status')
        if status == 'student':
            if v < 16 or v > 30:
                raise ValueError('Age must be between 16 and 30 for students')
        else:
            if v < 18 or v > 100:
                raise ValueError('Age must be between 18 and 100')
        return v
    
    @field_validator('student_course')
    @classmethod
    def validate_course(cls, v: Optional[int], info) -> Optional[int]:
        if info.data.get('status') == 'student' and (v is None or v < 1 or v > 4):
            raise ValueError('Course must be between 1 and 4 for students')
        return v
    
    @field_validator('student_education_level')
    @classmethod
    def validate_education_level(cls, v: Optional[str], info) -> Optional[str]:
        if info.data.get('status') == 'student':
            valid_levels = ['spo_9', 'spo_11', 'bachelor', 'master', 'postgraduate']
            if v not in valid_levels:
                raise ValueError(f'Education level must be one of {valid_levels} for students')
        return v

    @field_validator('teacher_experience')
    @classmethod
    def validate_teacher_exp(cls, v: Optional[int], info):
        if info.data.get('status') == 'teacher' and (v is None or v < 0 or v > 60):
            raise ValueError('Experience must be between 0 and 60 years')
        return v
    
    @field_validator('employee_experience')
    @classmethod
    def validate_employee_exp(cls, v: Optional[int], info):
        if info.data.get('status') == 'employee' and (v is None or v < 0 or v > 50):
            raise ValueError('Experience must be between 0 and 50 years')
        return v

@app.post("/api/user/profile")
async def save_user_profile(request: UserProfileRequest):
    """Save user profile information"""
    try:
        db = get_db()
        cur = db.cursor()

        try:
            cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS student_education_level VARCHAR(50)")
        except:
            pass

        cur.execute("SELECT id FROM users WHERE session_id = %s", (request.session_id,))
        existing = cur.fetchone()
        
        if existing:
            cur.execute("""
                UPDATE users 
                SET status = %s, age = %s, field = %s, field_other = %s,
                    student_course = %s, student_education_level = %s,
                    teacher_experience = %s, employee_experience = %s,
                    form_completed = 1, completed_at = CURRENT_TIMESTAMP
                WHERE session_id = %s
            """, (
                request.status, request.age, request.field, request.field_other,
                request.student_course, request.student_education_level,
                request.teacher_experience, request.employee_experience,
                request.session_id
            ))
            user_id = existing[0]
        else:
            cur.execute("""
                INSERT INTO users (session_id, status, age, field, field_other, 
                                  student_course, student_education_level,
                                  teacher_experience, employee_experience, form_completed)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
            """, (
                request.session_id, request.status, request.age, request.field, request.field_other,
                request.student_course, request.student_education_level,
                request.teacher_experience, request.employee_experience
            ))
            user_id = cur.lastrowid
        
        db.commit()

        if request.session_id in web_sessions:
            old_user_id = web_sessions[request.session_id].get("user_id")
            if old_user_id:
                cur.execute("""
                    UPDATE survey_results 
                    SET user_id = %s 
                    WHERE user_id = %s
                """, (user_id, old_user_id))
                db.commit()
        
        if request.session_id in web_sessions:
            web_sessions[request.session_id]["db_user_id"] = user_id

        db.close()
        
        return {"success": True, "user_id": user_id}
        
    except Exception as e:
        logging.error(f"Error saving user profile: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/user/profile/{session_id}")
async def get_user_profile(session_id: str):
    """Get user profile information"""
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        
        cur.execute("SELECT * FROM users WHERE session_id = %s", (session_id,))
        profile = cur.fetchone()
        db.close()
        
        if profile:
            if profile.get('completed_at'):
                profile['completed_at'] = profile['completed_at'].isoformat()
            return {"success": True, "profile": profile}
        else:
            return {"success": True, "profile": None}
            
    except Exception as e:
        logging.error(f"Error getting user profile: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


def get_user_id(session_id: str) -> int:
    """Get or create user_id for session"""
    if session_id not in web_sessions:
        import hashlib
        hash_object = hashlib.md5(session_id.encode())
        user_id = int(hash_object.hexdigest()[:8], 16) % 1000000

        web_sessions[session_id] = {
            "user_id": user_id,
            "chat_mode": False
        }
    return web_sessions[session_id]["user_id"]

@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the main HTML page"""
    with open("templates/website.html", "r", encoding="utf-8") as f:
        return f.read()


@app.post("/api/chat", response_model=ChatResponse)
async def chat_endpoint(request: ChatRequest):
    """Handle chat messages"""
    try:
        user_id = get_user_id(request.session_id)
        web_sessions[request.session_id]["chat_mode"] = True
        user_states[user_id] = "chat_mode"

        response = await bot_core.query_ollama(user_id, request.message)
        bot_core.update_history(user_id, request.message, response)
        
        return ChatResponse(
            response=response,
            success=True
        )
    except Exception as e:
        logging.error(f"Chat error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/chat/end/{session_id}")
async def end_chat(session_id: str):
    """End chat session"""
    if session_id in web_sessions:
        web_sessions[session_id]["chat_mode"] = False
        user_id = get_user_id(session_id)
        if user_id in user_states:
            user_states[user_id] = None
    return {"success": True}

@app.get("/api/advice/categories")
async def get_advice_categories():
    """Get all advice categories"""
    return {
        "breathing": "Техника дыхания 🌬️",
        "grounding": "Метод 5-4-3-2-1 🌿",
        "meditation": "Медитация 🧘",
        "massage": "Массаж 💆",
        "emergency": "Экспресс-помощь 🆘"
    }

@app.get("/api/advice/{category}")
async def get_advice(category: str, page: int = 1):
    """Get advice by category"""
    result = bot_core.get_advice_by_category(category, page)
    if not result:
        raise HTTPException(status_code=404, detail="Category not found")
    return result

@app.get("/api/advice/breathing/{index}")
async def get_breathing_technique(index: int):
    """Get specific breathing technique"""
    from features.advice import breathing
    if index < 0 or index >= len(breathing):
        raise HTTPException(status_code=404, detail="Technique not found")
    return {
        "index": index,
        "content": breathing[index]
    }

@app.post("/api/mood/entry")
async def add_mood_entry(request: MoodEntryRequest):
    """Add a mood entry"""
    try:
        user_id = get_user_id(request.session_id)
        
        if request.emoji not in MOODS:
            raise HTTPException(status_code=400, detail="Invalid emoji")
        
        entry = bot_core.add_mood_entry(
            user_id=user_id,
            emoji=request.emoji,
            mood_state=MOODS[request.emoji],
            note=request.note
        )
        return {"success": True, "entry": entry}
    except Exception as e:
        logging.error(f"Mood entry error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/mood/entries/{session_id}")
async def get_mood_entries(session_id: str, limit: int = 5, days: int = 7):
    """Get user's mood entries"""
    user_id = get_user_id(session_id)
    entries = bot_core.get_mood_entries(user_id, limit, days)
    return {"entries": entries}

@app.post("/api/mood/analyze/{session_id}")
async def analyze_mood(session_id: str):
    """Analyze user's mood entries"""
    user_id = get_user_id(session_id)
    entries = bot_core.get_mood_entries(user_id, days=7)
    
    if not entries:
        return {"analysis": "За последнюю неделю записей нет 🌱"}
    
    analysis = bot_core.analyze_mood_entries(entries)
    return {"analysis": analysis}

@app.post("/api/mood/advice/{session_id}")
async def get_mood_advice(session_id: str):
    """Get advice based on mood entries"""
    user_id = get_user_id(session_id)
    entries = bot_core.get_mood_entries(user_id, days=7)
    
    if not entries:
        return {"advice": "За последнюю неделю записей нет 🌱"}
    
    advice = bot_core.give_mood_advice(entries)
    return {"advice": advice}

@app.get("/api/history/{session_id}")
async def get_history(session_id: str, limit: int = 5):
    """Get chat history"""
    user_id = get_user_id(session_id)
    history = bot_core.get_history(user_id, limit)
    return history

@app.post("/api/history/clear/{session_id}")
async def clear_history(session_id: str):
    """Clear chat history"""
    user_id = get_user_id(session_id)
    bot_core.clear_history(user_id)
    return {"success": True}

@app.get("/api/reminders/types")
async def get_reminder_types():
    """Get available reminder types"""
    return REMINDER_TYPES

@app.get("/api/reminders/{session_id}")
async def get_reminders(session_id: str):
    """Get user's reminders"""
    user_id = get_user_id(session_id)
    reminders = bot_core.get_reminders(user_id)
    return {"reminders": reminders}

@app.post("/api/reminders/set")
async def set_reminder(request: ReminderRequest):
    """Set a reminder"""
    user_id = get_user_id(request.session_id)
    reminder = bot_core.set_reminder(
        user_id=user_id,
        reminder_type=request.reminder_type,
        time_str=request.time
    )
    return {"success": True, "reminder": reminder}

@app.post("/api/reminders/cancel/{session_id}/{reminder_type}")
async def cancel_reminder(session_id: str, reminder_type: str):
    """Cancel a reminder"""
    user_id = get_user_id(session_id)
    bot_core.cancel_reminder(user_id, reminder_type)
    return {"success": True}

from fastapi.responses import Response

@app.get("/api/mood/chart/{session_id}")
async def get_mood_chart(session_id: str):
    """Get mood statistics chart"""
    user_id = get_user_id(session_id)
    chart_bytes = bot_core.generate_mood_chart(user_id)
    
    if not chart_bytes:
        raise HTTPException(status_code=404, detail="No data for chart")
    
    return Response(content=chart_bytes, media_type="image/png")

@app.get("/api/reminders/check/{session_id}")
async def check_reminders(session_id: str):
    """Check for due reminders"""
    user_id = get_user_id(session_id)
    reminders = bot_core.get_reminders(user_id)
    
    current_time = datetime.now().strftime("%H:%M")
    due_reminders = []
    
    reminder_messages = {
        'diary': '🌸 Пора записать своё настроение в дневник!',
        'meditation': '🧘 Сделайте небольшую медитацию для расслабления',
        'drink_water': '💧 Не забудьте выпить воды!',
        'sleep_reminder': '😴 Пора готовиться ко сну',
        'stretch': '🤸 Сделайте небольшую разминку'
    }
    
    for reminder in reminders:
        if reminder['time'] == current_time and reminder.get('active', True):
            due_reminders.append({
                'type': reminder['type'],
                'message': reminder_messages.get(reminder['type'], 'Напоминание 🌿')
            })
    return {"due_reminders": due_reminders}

@app.post("/api/survey/save")
async def save_survey(request: SurveySaveRequest):
    """Save survey results to database"""
    try:
        user_id = get_user_id(request.session_id)
        
        bot_core.save_key_results(
            user_id=user_id,
            role=request.role,
            survey_key=request.survey_key,
            responses=request.responses
        )
        
        if request.open_ended_responses:
            db = get_db()
            cur = db.cursor()
            
            try:
                cur.execute("ALTER TABLE survey_results ADD COLUMN open_ended_responses TEXT")
            except:
                pass
            
            cur.execute("""
                UPDATE survey_results 
                SET open_ended_responses = %s
                WHERE user_id = %s AND survey_key = %s
                ORDER BY completed_at DESC LIMIT 1
            """, (json.dumps(request.open_ended_responses), user_id, request.survey_key))
            
            db.commit()
            db.close()
        
        return {"success": True}
    except Exception as e:
        logging.error(f"Save survey error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/admin/download-surveys")
async def download_surveys_as_excel(request: Request):
    """Download all survey results as Excel file (admin only)"""
    db = None
    try:
        body = await request.json()
        password = body.get("password", "")[:100]
        
        if not password or not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="Invalid password")
        
        db = get_db()
        cur = db.cursor(dictionary=True)
        
        cur.execute("""
            SELECT sr.id, sr.user_id, sr.role, sr.survey_key, sr.responses, sr.completed_at,
                sr.open_ended_responses, u.status, u.age, u.field, u.field_other, 
                u.student_course, u.student_education_level, u.teacher_experience, u.employee_experience
            FROM survey_results sr
            LEFT JOIN users u ON sr.user_id = u.id
            ORDER BY sr.completed_at DESC
            LIMIT 10000  -- Add reasonable limit
        """)
        
        results = cur.fetchall()
        
        if not results:
            raise HTTPException(status_code=404, detail="No survey results found")
        
        wb = Workbook()
        
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        survey_one_data = []
        survey_two_data = []
        survey_three_data = []
        
        for result in results:
            survey_key = result['survey_key']
            responses = json.loads(result['responses']) if isinstance(result['responses'], str) else result['responses']
            
            open_ended = None
            if result.get('open_ended_responses'):
                try:
                    open_ended = json.loads(result['open_ended_responses']) if isinstance(result['open_ended_responses'], str) else result['open_ended_responses']
                except:
                    open_ended = result['open_ended_responses']
            
            row = {
                'id': result['id'],
                'user_id': result['user_id'],
                'role': result['role'],
                'completed_at': result['completed_at'].strftime('%Y-%m-%d %H:%M:%S') if result['completed_at'] else 'N/A',
                'responses': responses,
                'responses_count': len(responses),
                'average_score': sum(responses) / len(responses) if responses else 0,
                'open_ended_responses': open_ended
            }
            
            if survey_key == 'surveyOne':
                survey_one_data.append(row)
            elif survey_key == 'surveyTwo':
                survey_two_data.append(row)
            elif survey_key == 'surveyThree':
                survey_three_data.append(row)
        
        if survey_one_data:
            ws1 = wb.create_sheet("Цифровая усталость")
            _create_survey_sheet(ws1, survey_one_data, "surveyOne")
        
        if survey_two_data:
            ws2 = wb.create_sheet("Шкала Ликерта")
            _create_survey_sheet(ws2, survey_two_data, "surveyTwo")
        
        if survey_three_data:
            ws3 = wb.create_sheet("Опросник Чена")
            _create_survey_sheet(ws3, survey_three_data, "surveyThree")
        
        summary_ws = wb.create_sheet("Сводка", 0)
        _create_summary_sheet(summary_ws, survey_one_data, survey_two_data, survey_three_data)
        
        await _create_profiles_sheet(wb, db)

        db.close()
        db = None
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=survey_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if db:
            db.close()

def _create_survey_sheet(ws, data, survey_key):
    """Helper function to create a survey sheet"""
    
    role_map = {
        'teacher': 'Преподаватель',
        'student': 'Студент',
        'employee': 'Сотрудник'
    }

    if survey_key == 'surveyTwo':
        headers = ['ID', 'User ID', 'Role', 'Completed At', 
                   'Responses (JSON)', 'Number of Answers', 'Average Score', 
                   'Open Ended Responses']
    else:
        headers = ['ID', 'User ID', 'Role', 'Completed At', 
                   'Responses (JSON)', 'Number of Answers', 'Average Score']
    
    if survey_key == 'surveyThree' and data:
        max_questions = max(len(d['responses']) for d in data)
        for i in range(1, max_questions + 1):
            headers.append(f'Q{i}')
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F5E83", end_color="4F5E83", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, record in enumerate(data, 2):
        col = 1
        ws.cell(row=row_idx, column=col, value=record['id']).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=record['user_id']).border = thin_border
        col += 1
        role_ru = role_map.get(record['role'], record['role'])
        ws.cell(row=row_idx, column=col, value=role_ru).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=record['completed_at']).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=str(record['responses'])).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=record['responses_count']).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=round(record['average_score'], 2)).border = thin_border
        col += 1
        
        if survey_key == 'surveyTwo':
            if record.get('open_ended_responses'):
                if isinstance(record['open_ended_responses'], list):
                    open_ended_text = ', '.join(str(item) for item in record['open_ended_responses'])
                else:
                    open_ended_text = str(record['open_ended_responses'])
            else:
                open_ended_text = ''
            ws.cell(row=row_idx, column=col, value=open_ended_text).border = thin_border
            col += 1

        if survey_key == 'surveyThree':
            for q_idx, answer in enumerate(record['responses'], 8):
                ws.cell(row=row_idx, column=q_idx, value=answer).border = thin_border
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'

def _create_summary_sheet(ws, survey_one_data, survey_two_data, survey_three_data):
    """Create summary sheet with statistics"""
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="Сводка по опросам")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    row = 3
    
    ws.cell(row=row, column=1, value="Всего заполненных опросов:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(survey_one_data) + len(survey_two_data) + len(survey_three_data))
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    ws.cell(row=row, column=1, value="Цифровая усталость (всего):").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(survey_one_data))
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    ws.cell(row=row, column=1, value="Шкала Ликерта (всего):").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(survey_two_data))
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    ws.cell(row=row, column=1, value="Опросник Чена (всего):").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(survey_three_data))
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    row += 2
    
    if survey_one_data:
        avg_one = round(sum(d['average_score'] for d in survey_one_data) / len(survey_one_data), 2)
        ws.cell(row=row, column=1, value="Средний балл (Цифровая усталость):").font = Font(bold=True)
        ws.cell(row=row, column=2, value=avg_one)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        row += 1
    
    if survey_two_data:
        avg_two = round(sum(d['average_score'] for d in survey_two_data) / len(survey_two_data), 2)
        ws.cell(row=row, column=1, value="Средний балл (Шкала Ликерта):").font = Font(bold=True)
        ws.cell(row=row, column=2, value=avg_two)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        row += 1
    
    if survey_three_data:
        avg_three = round(sum(d['average_score'] for d in survey_three_data) / len(survey_three_data), 2)
        ws.cell(row=row, column=1, value="Средний балл (Опросник Чена):").font = Font(bold=True)
        ws.cell(row=row, column=2, value=avg_three)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        row += 2
    
    ws.cell(row=row, column=1, value="Распределение по ролям:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).border = thin_border
    row += 1
    
    all_data = survey_one_data + survey_two_data + survey_three_data
    teachers = sum(1 for d in all_data if d['role'] == 'teacher')
    students = sum(1 for d in all_data if d['role'] == 'student')
    
    ws.cell(row=row, column=1, value="Преподаватели:")
    ws.cell(row=row, column=2, value=teachers)
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    ws.cell(row=row, column=1, value="Студенты:")
    ws.cell(row=row, column=2, value=students)
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

async def _create_profiles_sheet(wb, db):
    """Create a sheet with all user profiles"""
    cur = db.cursor(dictionary=True)
    
    cur.execute("""
        SELECT id, session_id, status, age, field, field_other, 
               student_course, student_education_level,
               teacher_experience, employee_experience, 
               completed_at, form_completed
        FROM users 
        WHERE form_completed = 1
        ORDER BY completed_at DESC
    """)
    
    users = cur.fetchall()
    cur.close()
    
    ws = wb.create_sheet("Профили участников")
    
    education_map = {
        'spo_9': 'СПО на базе 9 классов',
        'spo_11': 'СПО на базе 11 классов',
        'bachelor': 'Бакалавриат',
        'master': 'Магистратура',
        'postgraduate': 'Аспирантура'
    }
    
    headers = ['ID', 'Session ID', 'Статус', 'Возраст', 'Сфера деятельности', 
               'Другая сфера', 'Курс (студенты)', 'Уровень образования', 
               'Педагогический стаж', 'Опыт в образовании', 'Дата заполнения']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F5E83", end_color="4F5E83", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, user in enumerate(users, 2):
        col = 1
        ws.cell(row=row_idx, column=col, value=user['id']).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=user['session_id']).border = thin_border
        col += 1
        status_text = {'student': 'Студент', 'teacher': 'Преподаватель', 'employee': 'Сотрудник'}.get(user['status'], user['status'])
        ws.cell(row=row_idx, column=col, value=status_text).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=user['age']).border = thin_border
        col += 1
        
        field_text = {
            'it': 'ИТ-технологии', 'law': 'Юриспруденция', 'tourism': 'Туризм',
            'management': 'Менеджмент', 'psychology': 'Психология', 'medicine': 'Медицина',
            'economics': 'Экономика', 'logistics': 'Логистика', 'other': 'Другое'
        }.get(user['field'], user['field'])
        ws.cell(row=row_idx, column=col, value=field_text).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=user['field_other'] or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=user['student_course'] or '').border = thin_border
        col += 1
        
        edu_level = education_map.get(user['student_education_level'], user['student_education_level'] or '')
        ws.cell(row=row_idx, column=col, value=edu_level).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=user['teacher_experience'] or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=user['employee_experience'] or '').border = thin_border
        col += 1
        
        completed_at = user['completed_at'].strftime('%Y-%m-%d %H:%M:%S') if user['completed_at'] else ''
        ws.cell(row=row_idx, column=col, value=completed_at).border = thin_border
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 35)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'
    
    summary_ws = wb["Сводка"]
    row_num = summary_ws.max_row + 2
    summary_ws.cell(row=row_num, column=1, value="Всего заполнивших анкету:").font = Font(bold=True)
    summary_ws.cell(row=row_num, column=2, value=len(users))
    summary_ws.cell(row=row_num, column=1).border = thin_border
    summary_ws.cell(row=row_num, column=2).border = thin_border

@app.get("/admin")
async def admin_page():
    """Serve admin page"""
    with open("templates/admin.html", "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)